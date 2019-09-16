import requests
import openpyxl
import xlsxwriter
from PIL import Image

loc = "barcode generator.xlsx"  #enter the xcel sheet location here
book = openpyxl.load_workbook(loc)
sheet = book["Sheet1"]

workbook = xlsxwriter.Workbook('output_file.xlsx')
worksheet = workbook.add_worksheet()

for i in range(sheet.max_row):
   val = sheet.cell(row=i+1,column=1).value
    # url = "https://www.barcodesinc.com/generator/image.php?code="+val+"&style=197&type=C128B&width=115&height=50&xres=1&font=3"
   url=("https://barcode.tec-it.com/barcode.ashx?data="+str(val)+"&code=Code128&dpi=96&dataseparator=")
 #getting response    
   response = requests.get(url)
   name =str(val)+".png"
   with open(name,"wb") as fout:
         fout.write(response.content)
 #converting .png to .jpg        
   im = Image.open(name)
   rgb_im = im.convert('RGB')
   rgb_im.save(str(val)+"_jpg.jpg")
     

   worksheet.write(i,0, str(val))
   worksheet.insert_image(i,1,str(val)+"_jpg.jpg")

workbook.close()

